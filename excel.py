from openpyxl import load_workbook
import datetime
from function import format_date

wb = load_workbook(filename="res/final_pst.xlsx") # add your own file can't upload file to github

current_sheet = wb['Sheet1']
# Column 3 = Name
# Column 4 = FName
# Column 5 = DOB
# Column 6 = NICf



count = 1
for x in range(2, current_sheet.max_row+1):
  name = current_sheet.cell(row=x, column=3).value
  f_name = current_sheet.cell(row=x, column=4).value
  dob = current_sheet.cell(row=x, column=5).value
  nic = current_sheet.cell(row=x, column=6).value
  nic = str(nic)
  # print(datetime.datetime.strptime(dob, '%Y-%m-%d').strftime('%m/%d/%y'))
  # print(f'Name = {name}, Father Name = {f_name}, Date Of Birth = {dob}, NIC = {nic}, ')
  print(f'{format_date(dob)}')
  count += 1