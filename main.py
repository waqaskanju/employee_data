from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from function import fill_form, wait, format_date
from openpyxl import load_workbook

def open_employee():
  while True:
    try:
      myElem = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, 't_TreeNav_3')))
      myElem.click()
      print("t_TreeNav_3 is ready!")
      break
    except TimeoutException:
        print("Loading took too much time!")
  wait(10)
  while True:
    try:
      myElem = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#B23935025578417233 > span:nth-child(2)')))
      myElem.click()
      print("Employee Button is ready!")
      break
    except TimeoutException:
        print("Loading took too much time!")

driver = webdriver.Firefox(executable_path="./driver/geckodriver.exe")
driver.get("http://175.107.63.148:9090/ords/r/emis/human-resource-management-information-system-hrmis")
inputName = driver.find_element(By.ID, "P9999_USERNAME")
inputName.clear()
inputName.send_keys("")
inputPassword = driver.find_element(By.ID, "P9999_PASSWORD")
inputPassword.clear()
inputPassword.send_keys("")


btnSubmit = driver.find_element(By.ID, "B24597272248178128")
btnSubmit.click()
# wait(18)

while True:
  wait(5)
  try:
    wait(20)
    myElem = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, 't_Button_navControl')))
    myElem.send_keys(Keys.RETURN)
    print("t_Button_navControl is ready!")
    break
  except TimeoutException:
      print("Loading took too much time!")

wb = load_workbook(filename="res/final_pst.xlsx") # add your own file can't upload file to github

current_sheet = wb['Sheet1']
for x in range(2, 93):
  open_employee()
  wait(8)
  name = current_sheet.cell(row=x, column=3).value
  f_name = current_sheet.cell(row=x, column=4).value
  dob = current_sheet.cell(row=x, column=5).value
  nic = current_sheet.cell(row=x, column=6).value
  nic = str(nic)
  fill_form(name, f_name, format_date(dob), nic)
  current_sheet.cell(row = x, column = 10).value = 'PASSED'
  wb.save("res/final_pst.xlsx")
  wait(10)
  

